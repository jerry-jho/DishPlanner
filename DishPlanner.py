# -*- coding:utf-8 -*-
from PySide2.QtCore import Qt,QStandardPaths,QDir
from PySide2.QtWidgets import *
from QBootStrapWidget import QBootStrapTableWidget
import openpyxl

G_COL_TEXT = ['一', '二', '三', '四', '五', '六', '日']
G_ROW_TEXT = ['早', '午', '晚']
G_APPNAME = '食谱规划'
G_UNTITLED = '未命名'

class WSelector(QSplitter):
    def __init__(self, parent=None ,library_file = ''):
        QSplitter.__init__(self)

        self._library_file = library_file
        self.data = {}

        mainLay = QVBoxLayout()
        lay = QFormLayout()

        self.cbxCol = QComboBox()
        self.cbxCol.addItems(G_COL_TEXT)
        lay.addRow('选择日期',self.cbxCol)
        self.cbxRow = QComboBox()
        self.cbxRow.addItems(G_ROW_TEXT)       
        lay.addRow('选择时段',self.cbxRow)
        
        mainLay.addLayout(lay)
        self.lstContent = QListWidget()
        layContent = QVBoxLayout()

        self.btnRemove = QPushButton('删除当前项')
        layContent.addWidget(self.btnRemove)

        self.grpContent = QGroupBox('食谱')
        layContent.addWidget(self.lstContent)

        self.grpContent.setLayout(layContent)
        mainLay.addWidget(self.grpContent)

        layNew = QFormLayout()
        self.cbxType = QComboBox()
        
        self.cbxItem = QComboBox()
        self.edtAmount = QLineEdit()
        
        layNew.addRow('类别',self.cbxType)
        layNew.addRow('食物',self.cbxItem)
        layNew.addRow('数量',self.edtAmount)

        layContent.addLayout(layNew)

        self.btnAdd = QPushButton('添加到食谱')
        self.btnReloadLibrary = QPushButton('重新加载食品库')
        self.btnLoadData = QPushButton('加载食谱')
        self.btnSaveData = QPushButton('保存食谱')
        layContent.addWidget(self.btnAdd)
        layContent.addWidget(self.btnReloadLibrary)
        layContent.addWidget(self.btnLoadData)
        layContent.addWidget(self.btnSaveData)
        
        self.grpPreview = QGroupBox('预览')
        self.layPreview = QVBoxLayout()
        self.tblPreview = QBootStrapTableWidget(len(G_ROW_TEXT),len(G_COL_TEXT))
        self.tblPreview.setHorizontalHeaderLabels(G_COL_TEXT)
        self.tblPreview.setVerticalHeaderLabels(G_ROW_TEXT)

        self.layPreview.addWidget(self.tblPreview)
        self.grpPreview.setLayout(self.layPreview)

        self.grpSettings = QGroupBox('设置')
        self.grpSettings.setLayout(mainLay)

        self.addWidget(self.grpSettings)
        self.addWidget(self.grpPreview)
        self.setStretchFactor(1, 4)
        self.resize(1200,400)

        self.loadLibrary()
        self.btnReloadLibrary.clicked.connect(self.loadLibrary)

        self.updateItem()
        self.cbxType.currentTextChanged.connect(self.updateItem)
        self.btnAdd.clicked.connect(self.addItem)
        self.btnRemove.clicked.connect(self.removeItem)
        self.cbxCol.currentIndexChanged.connect(self.updateListAndPreview)
        self.cbxRow.currentIndexChanged.connect(self.updateListAndPreview)
        self.btnLoadData.clicked.connect(self.loadData)
        self.btnSaveData.clicked.connect(self.saveData)

        self.docPath = None
        self.setDocModified(False)

    def setDocModified(self,stat):
        self.docModified = stat
        if stat:
            if self.docPath is None:
                self.setWindowTitle(G_APPNAME + ' - [' + G_UNTITLED + '*]')
            else:
                self.setWindowTitle(G_APPNAME + ' - [' + self.docPath + '*]')
        else:
            if self.docPath is None:
                self.setWindowTitle(G_APPNAME + ' - [' + G_UNTITLED + ']')
            else:
                self.setWindowTitle(G_APPNAME + ' - [' + self.docPath + ']')                        

    def removeItem(self):
        key = (self.cbxRow.currentIndex(),self.cbxCol.currentIndex())
        lst = self.data.get(key,[])
        try:
            to_remove = str(self.lstContent.currentItem().text())
            lst.remove(to_remove)
        except:
            return
        self.data[key] = lst
        self.updateListAndPreview()
        self.setDocModified(True)

    def updateItem(self):
        lst = self.db.get(str(self.cbxType.currentText()),[])
        self.cbxItem.clear()
        self.cbxItem.addItems(lst)

    def addItem(self):
        txt = str(self.cbxItem.currentText())
        val = str(self.edtAmount.text())
        if val != '':
            txt += '[' + val + ']'
        key = (self.cbxRow.currentIndex(),self.cbxCol.currentIndex())
        if not key in self.data.keys():
            self.data[key] = []
        self.data[key].append(txt)
        self.updateListAndPreview()
        self.setDocModified(True)

    def updateListAndPreview(self):
        key = (self.cbxRow.currentIndex(),self.cbxCol.currentIndex())
        lst = self.data.get(key,[])
        self.lstContent.clear()
        self.lstContent.addItems(lst)
        self.tblPreview.setText(self.cbxRow.currentIndex(),self.cbxCol.currentIndex(),'\n'.join(lst))
        

    def loadLibrary(self):
        wb = openpyxl.load_workbook(self._library_file, read_only=True)
        sht = wb.worksheets[0]
        db = OrderedDict()
        for row in sht.rows:
            for i,cell in enumerate(row):
                if i == 0:
                    key = cell.value
                    db[key] = []
                elif cell.value is not None:
                    db[key].append(cell.value)
        wb.close()
        self.db = db
        self.cbxType.clear()
        self.cbxType.addItems(self.db.keys())
        self.updateItem()

    def loadData(self):
        if self.docModified:
            rtn = QMessageBox.question(self,G_APPNAME,
                '当前食谱没有保存，是否继续打开新的食谱',QMessageBox.Yes | QMessageBox.No,QMessageBox.No)
            if rtn != QMessageBox.Yes:
                return
        
        path,ftype = QFileDialog.getOpenFileName(self,'打开食谱',QStandardPaths.writableLocation(QStandardPaths.DocumentsLocation),'Excel 文件(*.xlsx)')
        if path != '' and os.path.exists(path):
            try:
                wb = openpyxl.load_workbook(path, read_only=True)
            except:
                QMessageBox.critical(self,G_APPNAME,'%s 不是合法的数据文件' % path)
                return
            try:
                sht = wb.worksheets[0]
            except:
                QMessageBox.critical(self,G_APPNAME,'%s 没有合法的工作表' % path)
                return
            self.data.clear()
            for r,row in enumerate(sht.rows):
                if r > 3:
                    break
                for c,cell in enumerate(row):
                    if c > 7:
                        break
                    key = cell.value
                    if key is not None:
                        self.data[(r,c)] = key.split('\n')
                        self.tblPreview.setText(r,c,key)
            self.updateListAndPreview()
            self.docPath = path
            self.setDocModified(False)
            wb.close()
    def saveData(self):
        if not self.docModified:
            return
        if self.docPath is None:
            path,ftype = QFileDialog.getSaveFileName(self,'保存食谱',QStandardPaths.writableLocation(QStandardPaths.DocumentsLocation),'Excel 文件(*.xlsx)')
            if path == '':
                return
            if QDir(path) == QDir(self._library_file):
                QMessageBox.critical(self,G_APPNAME,'%s 是库文件，不能用于保存食谱' % path)
                return
        else:                
            path = self.docPath
        wb = openpyxl.Workbook()
        ws = wb.active
        for k,v in self.data.items():
            cell = ws.cell(row=k[0]+1,column=k[1]+1)
            cell.value = "\n".join(v)
        try:
            wb.save(path)
            wb.close()
        except:
            QMessageBox.critical(self,G_APPNAME,'%s 保存失败' % path)
            return
        self.docPath = path
        self.setDocModified(False)

if __name__ == '__main__':
    app = QApplication([])

    #library
    
    import os
    from collections import OrderedDict

    #print(db)
    w = WSelector(library_file=os.path.realpath(os.path.dirname(__file__)) + '/library.xlsx')
    w.show()
    app.exec_()